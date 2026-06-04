"""Convert the MMA synthetic data workbook into the customer JSON used by app.py.

The app expects a flat list of customer dictionaries with keys such as:
`id`, `name`, `credit_score`, `transactions`, `loans`, and `risk_tier`.

This converter preserves every column from the Excel `synthetic_data` sheet at
the top level of each customer object, then adds the app-compatible fields that
the dashboard already reads.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


INPUT_XLSX = Path("MMA - Practicum Synthetic Data.xlsx")
OUTPUT_JSON = Path("data/customers.json")
REQUESTED_SHEET = "synthetic_data"
SHEET_FALLBACKS = {
    "synthetic_data": ["Synethic_Data", "Synthetic_Data"],
}

EMPLOYMENT_JOBS = {
    "Full-Time": [
        "Relationship Manager",
        "Financial Analyst",
        "Operations Specialist",
        "Project Manager",
        "Registered Nurse",
        "Teacher",
        "Software Engineer",
        "Account Manager",
    ],
    "Part-Time": [
        "Retail Associate",
        "Teaching Assistant",
        "Barista",
        "Customer Support Representative",
    ],
    "Self-Employed": [
        "Consultant",
        "Freelancer",
        "Business Owner",
        "Independent Contractor",
    ],
    "Retired": ["Retired"],
}

RISK_BASE = {
    "Primacy": 0.24,
    "Near Primacy": 0.38,
    "Non-Primacy": 0.56,
}

BOOL_PREFIXES = ("has_", "is_")
BOOL_EXACT = {
    "primacy_flag",
    "product_usage_flag",
    "digital_engagement_flag_30days",
    "unsolicited_order_flag_ip",
    "unsolicited_order_flag_gic",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=INPUT_XLSX, help="Path to the Excel workbook")
    parser.add_argument("--output", type=Path, default=OUTPUT_JSON, help="Path to the output customers.json")
    parser.add_argument("--sheet", default=REQUESTED_SHEET, help="Workbook sheet to convert")
    parser.add_argument("--limit", type=int, default=None, help="Only convert the first N customer rows")
    parser.add_argument("--dry-run", action="store_true", help="Validate conversion without writing the JSON file")
    return parser.parse_args()


def print_error(message: str) -> None:
    print(f"ERROR: {message}", file=sys.stderr)


def print_warning(message: str) -> None:
    print(f"WARNING: {message}")


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def stable_hash(customer_id: str, salt: str) -> int:
    digest = hashlib.sha256(f"{customer_id}|{salt}".encode("utf-8")).hexdigest()
    return int(digest[:16], 16)


def seeded_int(customer_id: str, salt: str, low: int, high: int) -> int:
    return low + stable_hash(customer_id, salt) % (high - low + 1)


def seeded_float(customer_id: str, salt: str, low: float, high: float) -> float:
    span = high - low
    return low + (stable_hash(customer_id, salt) / float(16**16 - 1)) * span


def seeded_choice(customer_id: str, salt: str, options: list[str]) -> str:
    return options[stable_hash(customer_id, salt) % len(options)]


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and value != value:  # NaN check (pandas empty cells)
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def clean_text(value: Any, default: str = "") -> str:
    if is_missing(value):
        return default
    text = str(value).replace("\n", ", ").strip()
    return " ".join(text.split()) or default


def as_float(value: Any, default: float = 0.0) -> float:
    if is_missing(value):
        return default
    if isinstance(value, bool):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def as_int(value: Any, default: int = 0) -> int:
    return int(round(as_float(value, default)))


def parse_bool(value: Any) -> bool | None:
    if is_missing(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value in (0, 1):
            return bool(value)
        return None
    text = str(value).strip().lower()
    if text in {"true", "t", "yes", "y", "1", "active", "open"}:
        return True
    if text in {"false", "f", "no", "n", "0", "inactive", "closed"}:
        return False
    return None


def _parse_date_str(text: str) -> str | None:
    """Parse common date string formats (M/D/YYYY, YYYY-MM-DD, etc.) into ISO format."""
    text = text.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text or None


def iso_date(value: Any) -> str | None:
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return _parse_date_str(value)
    return None


def clean_cell(column: str, value: Any) -> Any:
    if is_missing(value):
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and any(c in value for c in ("/"," -")):
        # Try to parse date-like strings that came from CSV
        parsed = _parse_date_str(value)
        if parsed and parsed != value:
            return parsed
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return float(value)
    if isinstance(value, int):
        return value

    column_lower = column.lower()
    if column_lower.startswith(BOOL_PREFIXES) or column_lower.endswith("_flag") or column_lower in BOOL_EXACT:
        parsed = parse_bool(value)
        if parsed is not None:
            return parsed

    return clean_text(value)


def resolve_sheet_name(workbook: Any, requested_sheet: str) -> str:
    sheetnames = list(workbook.sheetnames)
    if requested_sheet in sheetnames:
        return requested_sheet

    requested_lower = requested_sheet.lower()
    for sheet in sheetnames:
        if sheet.lower() == requested_lower:
            return sheet

    for candidate in SHEET_FALLBACKS.get(requested_lower, []):
        if candidate in sheetnames:
            return candidate

    raise KeyError(
        f"Sheet '{requested_sheet}' not found. Available sheets: {', '.join(sheetnames)}"
    )


def parse_goal_summary(summary: Any) -> list[dict[str, Any]]:
    text = clean_text(summary)
    if not text:
        return []

    goals: list[dict[str, Any]] = []
    for part in text.split(" | "):
        pieces = [piece.strip() for piece in part.split(":")]
        if len(pieces) < 4:
            continue
        goals.append(
            {
                "purpose": pieces[0],
                "status": pieces[1],
                "target_amount": round(as_float(pieces[2], 0.0), 2),
                "target_date": pieces[3],
            }
        )
    return goals


def derive_age(row: dict[str, Any], customer_id: str) -> int:
    retirement_age = as_int(row.get("retirement_age"), 0)
    annual_income = as_int(row.get("annual_income"), 0)
    goal_purpose = clean_text(row.get("goal_purpose"), "General Savings").lower()

    if retirement_age:
        base = retirement_age - seeded_int(customer_id, "retirement_gap", 7, 24)
    elif "retirement" in goal_purpose:
        base = seeded_int(customer_id, "goal_retirement_age", 45, 63)
    elif annual_income >= 130_000:
        base = seeded_int(customer_id, "income_high_age", 32, 58)
    elif annual_income >= 80_000:
        base = seeded_int(customer_id, "income_mid_age", 27, 52)
    else:
        base = seeded_int(customer_id, "income_low_age", 22, 46)
    return int(clamp(base, 20, 78))


def derive_employment(row: dict[str, Any], customer_id: str, age: int) -> tuple[str, str, str, int]:
    annual_income = as_int(row.get("annual_income"), 0)
    primary_segment = clean_text(row.get("primary_segment"), "Non-Primacy")

    if age >= 65:
        status = "Retired"
    elif annual_income >= 135_000 and stable_hash(customer_id, "self_employed_flag") % 5 == 0:
        status = "Self-Employed"
    elif annual_income < 40_000:
        status = "Part-Time"
    else:
        status = "Full-Time"

    job_title = seeded_choice(customer_id, "job_title", EMPLOYMENT_JOBS[status])
    if status == "Retired":
        employer = "-"
        years = 0
    elif status == "Self-Employed":
        employer = f"{clean_text(row.get('customer_name'), 'Client').split()[0]} Advisory"
        years = seeded_int(customer_id, "self_employed_years", 1, min(age - 18, 20))
    else:
        employer = f"{primary_segment} Client Services"
        years = seeded_int(customer_id, "employer_years", 0, min(age - 18, 18))
    return status, job_title, employer, max(years, 0)


def derive_member_since(row: dict[str, Any], customer_id: str) -> str:
    business_effective_date = row.get("business_effective_date")
    if isinstance(business_effective_date, str):
        try:
            start_date = datetime.fromisoformat(business_effective_date).date()
        except ValueError:
            start_date = date.today()
    elif isinstance(business_effective_date, date):
        start_date = business_effective_date
    else:
        start_date = date.today()

    offset_days = seeded_int(customer_id, "member_since_days", 180, 365 * 12)
    return (start_date - timedelta(days=offset_days)).isoformat()


def derive_balances(row: dict[str, Any]) -> tuple[int, int, int, int]:
    checking = max(0, round(as_float(row.get("balance_cad_1month_bb_d2d"), 0.0)))
    savings = max(0, round(as_float(row.get("balance_cad_1month_hisa"), 0.0)))
    investment = max(
        0,
        round(
            max(
                as_float(row.get("gic_market_value_cad_gic"), 0.0),
                as_float(row.get("gic_account_amount_cad_gic"), 0.0),
                as_float(row.get("gic_face_value_cad_gic"), 0.0),
                as_float(row.get("asset_amount_cad_hisa"), 0.0) - savings,
            )
        ),
    )
    total_assets = checking + savings + investment
    return checking, savings, investment, total_assets


def derive_monthly_expenses(row: dict[str, Any], customer_id: str, annual_income: int) -> int:
    observed = sum(
        as_float(row.get(key), 0.0)
        for key in (
            "net_grocery_amt_30days",
            "net_dining_amt_30days",
            "net_fuel_amt_30days",
            "net_travel_amt_30days",
            "net_daily_transit_amt_30days",
            "net_health_amt_30days",
            "net_telecom_utilities_amt_30days",
            "net_recurring_payment_amt_30days",
        )
    )
    baseline_ratio = seeded_float(customer_id, "expense_ratio", 0.42, 0.74)
    baseline = annual_income / 12 * baseline_ratio
    return int(round(clamp(max(observed, baseline), 500, 12_000)))


def derive_open_account_count(row: dict[str, Any]) -> int:
    count = 0
    for key, value in row.items():
        if key.startswith("has_open_") or key.startswith("has_active_"):
            parsed = parse_bool(value)
            if parsed is True:
                count += 1
    return max(count, 1)


def derive_total_debt(
    row: dict[str, Any],
    customer_id: str,
    annual_income: int,
    total_assets: int,
    monthly_expenses: int,
    goals: list[dict[str, Any]],
) -> tuple[int, float]:
    primary_segment = clean_text(row.get("primary_segment"), "Non-Primacy")
    recurring = as_float(row.get("net_recurring_payment_amt_30days"), 0.0)
    liabilities = max(
        as_float(row.get("liability_amount_cad_hisa"), 0.0),
        as_float(row.get("target_spending_amount"), 0.0),
    )
    debt_goal_amount = sum(g["target_amount"] for g in goals if "debt" in g["purpose"].lower())
    asset_relief = min(0.18, total_assets / max(annual_income, 1) * 0.08)
    recurring_burden = min(0.14, recurring / max(annual_income / 12, 1) * 0.30)
    goal_pressure = 0.18 if debt_goal_amount else 0.0
    noise = seeded_float(customer_id, "dti_noise", -0.05, 0.08)
    dti_ratio = clamp(
        RISK_BASE.get(primary_segment, 0.48) + recurring_burden + goal_pressure - asset_relief + noise,
        0.08,
        1.25,
    )
    debt_floor = max(recurring * 18, monthly_expenses * 3, debt_goal_amount * 0.45, liabilities * 0.4)
    total_debt = int(round(max(annual_income * dti_ratio, debt_floor)))
    return total_debt, round(total_debt / max(annual_income, 1), 4)


def derive_late_payments(customer_id: str, primary_segment: str, debt_to_income_ratio: float) -> int:
    if debt_to_income_ratio < 0.28 and primary_segment == "Primacy":
        return 0
    if debt_to_income_ratio < 0.42:
        return seeded_int(customer_id, "late_low", 0, 1)
    if debt_to_income_ratio < 0.70:
        return seeded_int(customer_id, "late_mid", 1, 3)
    return seeded_int(customer_id, "late_high", 2, 5)


def derive_credit_score(
    row: dict[str, Any],
    customer_id: str,
    annual_income: int,
    total_assets: int,
    debt_to_income_ratio: float,
    num_late_payments: int,
    bankruptcy_history: bool,
) -> int:
    primary_segment = clean_text(row.get("primary_segment"), "Non-Primacy")
    digital = parse_bool(row.get("has_digital_engagement_last_30days"))
    if digital is None:
        digital = parse_bool(row.get("digital_engagement_flag_30days"))

    score = 690.0
    score += 35 if primary_segment == "Primacy" else 8 if primary_segment == "Near Primacy" else -28
    score += clamp((annual_income - 60_000) / 4_000, -40, 45)
    score += clamp((total_assets - annual_income * 0.25) / 5_000, -35, 55)
    score -= clamp(debt_to_income_ratio * 170, 0, 230)
    score -= num_late_payments * 18
    score -= 90 if bankruptcy_history else 0
    score += 12 if digital is True else 0
    score += seeded_float(customer_id, "credit_noise", -22, 22)
    return int(round(clamp(score, 300, 850)))


def derive_risk(customer: dict[str, Any]) -> tuple[str, float]:
    score = 0.0
    score += (customer["credit_score"] - 300) / 550 * 40
    score += max(0.0, (1 - customer["debt_to_income_ratio"])) * 20
    score += (
        15.0
        if customer["employment_status"] == "Full-Time"
        else 9.0
        if customer["employment_status"] in ("Self-Employed", "Retired")
        else 5.0
    )
    score += 0.0 if customer["bankruptcy_history"] else 10.0
    score += clamp((customer["annual_income"] - 20_000) / 180_000 * 15, 0, 15)
    score = round(clamp(score, 0, 100), 1)
    risk_tier = "Low" if score >= 72 else "Medium" if score >= 48 else "High"
    return risk_tier, score


def amortized_payment(amount: float, annual_rate: float, months: int) -> float:
    monthly_rate = annual_rate / 100 / 12
    if amount <= 0:
        return 0.0
    if monthly_rate == 0:
        return round(amount / months, 2)
    payment = amount * monthly_rate / (1 - (1 + monthly_rate) ** -months)
    return round(payment, 2)


def derive_loans(
    row: dict[str, Any],
    customer_id: str,
    total_debt: int,
    credit_score: int,
    num_late_payments: int,
    business_date: str,
) -> list[dict[str, Any]]:
    if total_debt < 5_000:
        return []

    debt_goal = "debt" in clean_text(row.get("financial_goals_summary")).lower()
    if debt_goal:
        loan_type, months = "Line of Credit", 24
    elif total_debt > 90_000:
        loan_type, months = "Mortgage", 300
    elif total_debt > 30_000:
        loan_type, months = "Auto", 72
    else:
        loan_type, months = "Personal", 60

    rate = clamp(
        4.2 + (760 - credit_score) / 45 + seeded_float(customer_id, "rate_noise", -0.6, 1.0),
        3.5,
        18.9,
    )
    primary_amount = int(round(total_debt * 0.7))
    start = datetime.fromisoformat(business_date).date() - timedelta(days=seeded_int(customer_id, "loan_start", 120, 1800))
    end = start + timedelta(days=30 * months)
    loans = [
        {
            "type": loan_type,
            "amount": primary_amount,
            "rate": round(rate, 2),
            "status": "Active",
            "monthly_payment": amortized_payment(primary_amount, rate, months),
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "missed_payments": min(num_late_payments, 3),
        }
    ]

    if total_debt - primary_amount >= 8_000:
        closed_amount = total_debt - primary_amount
        closed_start = start - timedelta(days=30 * 36)
        closed_end = closed_start + timedelta(days=30 * 36)
        closed_rate = clamp(rate + 0.8, 3.5, 18.9)
        loans.append(
            {
                "type": "Personal",
                "amount": int(round(closed_amount)),
                "rate": round(closed_rate, 2),
                "status": "Closed",
                "monthly_payment": amortized_payment(closed_amount, closed_rate, 36),
                "start_date": closed_start.isoformat(),
                "end_date": closed_end.isoformat(),
                "missed_payments": min(num_late_payments, 1),
            }
        )
    return loans


def build_transactions(row: dict[str, Any], business_date: str) -> list[dict[str, Any]]:
    transactions: list[dict[str, Any]] = []

    transaction_specs = [
        ("effective_date_dda", "payment_category_dda", "payment_subcategory_dda", "transaction_amount_dda", "DDA"),
        ("effective_date_bb", "payment_category_bb", "payment_subcategory_bb", "transaction_amount_bb", "BB"),
        ("transaction_date_mft", "payment_category_mft", "payment_subcategory_mft", "transaction_amount_cad_mft", "MFT"),
        ("transaction_date_gic", "payment_category_gic", "payment_subcategory_gic", "transaction_amount_gic", "GIC"),
    ]

    for date_key, category_key, subcategory_key, amount_key, fallback_label in transaction_specs:
        amount_raw = row.get(amount_key)
        if is_missing(amount_raw):
            continue
        amount = as_float(amount_raw, 0.0)
        if amount == 0:
            continue
        txn_date = iso_date(row.get(date_key)) or business_date
        category = clean_text(row.get(category_key), fallback_label)
        subcategory = clean_text(row.get(subcategory_key))
        description = f"{category} - {subcategory}" if subcategory else f"{category} - Imported"
        transactions.append(
            {
                "date": txn_date,
                "category": category,
                "amount": round(amount, 2),
                "description": description,
            }
        )

    summary_specs = [
        (0, "Payroll", as_float(row.get("deposit_amount_cad_1month_bb_d2d"), 0.0), "Payroll - Direct Deposit"),
        (2, "Transfer", as_float(row.get("deposit_amount_cad_1month_hisa"), 0.0), "Transfer - Savings Deposit"),
        (4, "Groceries", -as_float(row.get("net_grocery_amt_30days"), 0.0), "Groceries - Card Activity"),
        (6, "Dining", -as_float(row.get("net_dining_amt_30days"), 0.0), "Dining - Card Activity"),
        (8, "Transport", -as_float(row.get("net_fuel_amt_30days"), 0.0), "Transport - Fuel Spend"),
        (10, "Travel", -as_float(row.get("net_travel_amt_30days"), 0.0), "Travel - Card Activity"),
        (12, "Transport", -as_float(row.get("net_daily_transit_amt_30days"), 0.0), "Transport - Transit Spend"),
        (14, "Healthcare", -as_float(row.get("net_health_amt_30days"), 0.0), "Healthcare - Card Activity"),
        (16, "Utilities", -as_float(row.get("net_telecom_utilities_amt_30days"), 0.0), "Utilities - Telecom and Utilities"),
        (18, "Insurance", -as_float(row.get("net_recurring_payment_amt_30days"), 0.0), "Insurance - Recurring Payments"),
    ]

    anchor = datetime.fromisoformat(business_date).date()
    for day_offset, category, amount, description in summary_specs:
        if abs(amount) < 0.01:
            continue
        transactions.append(
            {
                "date": (anchor - timedelta(days=day_offset)).isoformat(),
                "category": category,
                "amount": round(amount, 2),
                "description": description,
            }
        )

    unique: dict[tuple[str, str, float, str], dict[str, Any]] = {}
    for txn in transactions:
        key = (txn["date"], txn["category"], txn["amount"], txn["description"])
        unique[key] = txn
    deduped = list(unique.values())
    deduped.sort(key=lambda item: item["date"], reverse=True)
    return deduped[:50]


def load_excel_rows(input_path: Path, sheet_name: str, limit: int | None = None) -> tuple[list[dict[str, Any]], str]:
    if not input_path.exists():
        raise FileNotFoundError(f"Excel file not found: {input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    actual_sheet_name = resolve_sheet_name(workbook, sheet_name)
    worksheet = workbook[actual_sheet_name]

    row_iter = worksheet.iter_rows(values_only=True)
    header = [clean_text(value) for value in next(row_iter)]
    rows: list[dict[str, Any]] = []

    for excel_row_number, values in enumerate(row_iter, start=2):
        if limit is not None and len(rows) >= limit:
            break
        raw_row = {
            column: clean_cell(column, value)
            for column, value in zip(header, values)
            if column
        }

        customer_id = clean_text(raw_row.get("customer_id"))
        customer_name = clean_text(raw_row.get("customer_name"))
        if not customer_id or not customer_name:
            print_warning(
                f"Row {excel_row_number} is missing customer_id or customer_name and will be skipped."
            )
            continue
        rows.append(raw_row)

    return rows, actual_sheet_name


def convert_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    customer = dict(raw_row)

    customer_id = clean_text(raw_row.get("customer_id"))
    customer_name = clean_text(raw_row.get("customer_name"), customer_id)
    annual_income = as_int(raw_row.get("annual_income"), 0)
    goals = parse_goal_summary(raw_row.get("financial_goals_summary"))
    age = derive_age(raw_row, customer_id)
    employment_status, job_title, employer, years_at_employer = derive_employment(raw_row, customer_id, age)
    member_since = derive_member_since(raw_row, customer_id)
    checking_balance, savings_balance, investment_balance, total_assets = derive_balances(raw_row)
    monthly_expenses = derive_monthly_expenses(raw_row, customer_id, annual_income)
    total_debt, debt_to_income_ratio = derive_total_debt(
        raw_row,
        customer_id,
        annual_income,
        total_assets,
        monthly_expenses,
        goals,
    )

    primary_segment = clean_text(raw_row.get("primary_segment"), "Non-Primacy")
    num_late_payments = derive_late_payments(customer_id, primary_segment, debt_to_income_ratio)
    bankruptcy_history = (
        debt_to_income_ratio > 0.95
        and primary_segment == "Non-Primacy"
        and stable_hash(customer_id, "bankruptcy") % 7 == 0
    )
    credit_score = derive_credit_score(
        raw_row,
        customer_id,
        annual_income,
        total_assets,
        debt_to_income_ratio,
        num_late_payments,
        bankruptcy_history,
    )
    business_effective_date = iso_date(raw_row.get("business_effective_date")) or date.today().isoformat()

    customer.update(
        {
            "id": customer_id,
            "name": customer_name,
            "age": age,
            "gender": clean_text(raw_row.get("gender"), "Unspecified"),
            "email": clean_text(raw_row.get("email"), f"{customer_id.lower()}@example.com"),
            "phone": clean_text(
                raw_row.get("phone"),
                f"(555) {seeded_int(customer_id, 'phone_mid', 100, 999):03d}-{seeded_int(customer_id, 'phone_last', 0, 9999):04d}",
            ),
            "address": clean_text(raw_row.get("address"), "Address not provided"),
            "member_since": member_since,
            "employment_status": employment_status,
            "job_title": clean_text(raw_row.get("job_title"), job_title),
            "employer": clean_text(raw_row.get("employer"), employer),
            "years_at_employer": as_int(raw_row.get("years_at_employer"), years_at_employer),
            "annual_income": annual_income,
            "monthly_expenses": monthly_expenses,
            "credit_score": credit_score,
            "num_late_payments": as_int(raw_row.get("num_late_payments"), num_late_payments),
            "months_since_last_delinquency": (
                None
                if as_int(raw_row.get("num_late_payments"), num_late_payments) == 0
                else as_int(raw_row.get("months_since_last_delinquency"), seeded_int(customer_id, "months_since_delinquency", 1, 48))
            ),
            "total_debt": total_debt,
            "debt_to_income_ratio": debt_to_income_ratio,
            "payment_to_income_ratio": round((total_debt * 0.015) / max(annual_income / 12, 1), 4) if annual_income else 0.0,
            "num_open_accounts": as_int(raw_row.get("num_open_accounts"), derive_open_account_count(raw_row)),
            "bankruptcy_history": parse_bool(raw_row.get("bankruptcy_history"))
            if parse_bool(raw_row.get("bankruptcy_history")) is not None
            else bankruptcy_history,
            "checking_balance": checking_balance,
            "savings_balance": savings_balance,
            "investment_balance": investment_balance,
            "total_assets": total_assets,
            "financial_goals": goals,
            "business_effective_date": business_effective_date,
            "customer_id": customer_id,
            "customer_name": customer_name,
        }
    )

    customer["loans"] = derive_loans(
        customer,
        customer_id,
        customer["total_debt"],
        customer["credit_score"],
        customer["num_late_payments"],
        business_effective_date,
    )
    customer["transactions"] = build_transactions(customer, business_effective_date)
    customer["risk_tier"], customer["risk_score"] = derive_risk(customer)
    return customer


def main() -> int:
    args = parse_args()

    try:
        rows, actual_sheet_name = load_excel_rows(args.input, args.sheet, args.limit)
    except FileNotFoundError as exc:
        print_error(str(exc))
        return 1
    except KeyError as exc:
        print_error(str(exc))
        return 1
    except Exception as exc:
        print_error(f"Failed to load Excel workbook: {exc}")
        return 1

    print(f"Excel file loaded successfully: {args.input}")
    if actual_sheet_name == args.sheet:
        print(f"{args.sheet} sheet loaded successfully")
    else:
        print(f"{args.sheet} sheet loaded successfully using workbook sheet '{actual_sheet_name}'")

    customers = [convert_row(row) for row in rows]
    print(f"Number of customers converted: {len(customers)}")

    if args.dry_run:
        print("Dry run complete. customers.json was not written.")
        if customers:
            print(json.dumps(customers[0], indent=2, ensure_ascii=False)[:4000])
        return 0

    try:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        with args.output.open("w", encoding="utf-8") as handle:
            json.dump(customers, handle, indent=2, ensure_ascii=False)
    except Exception as exc:
        print_error(f"Failed to save customers.json: {exc}")
        return 1

    print(f"customers.json saved successfully: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
